#!/usr/bin/env python3
"""完善并统一「7天全预算」「行前准备与携带清单」格式，与现有工作簿风格一致。"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

EXCEL = Path.home() / 'Desktop' / 'KL_Travel_Guide_2026-07-12_to_15.xlsx'

# ── 配色（与现有表一致）──
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=14)
SUB_FONT = Font(name='微软雅黑', color='44546A', size=9, italic=True)
BODY_FONT = Font(name='微软雅黑', size=9)
BODY_BOLD = Font(name='微软雅黑', size=9, bold=True)
TOTAL_FILL = PatternFill('solid', fgColor='2F5496')
TOTAL_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
ALT_FILL = PatternFill('solid', fgColor='F7F9FC')
THIN = Side(style='thin', color='D0D7E2')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

CAT_FILL = {
    '大交通': PatternFill('solid', fgColor='DCE6F1'),
    '住宿': PatternFill('solid', fgColor='E2EFDA'),
    '餐饮': PatternFill('solid', fgColor='FCE4D6'),
    '当地交通': PatternFill('solid', fgColor='DEEBF7'),
    '游玩体验': PatternFill('solid', fgColor='E4DFEC'),
    '手续杂项': PatternFill('solid', fgColor='F2F2F2'),
    '旅行用品': PatternFill('solid', fgColor='FFF2CC'),
    '证件': PatternFill('solid', fgColor='DCE6F1'),
    '衣物': PatternFill('solid', fgColor='E2EFDA'),
    '防晒': PatternFill('solid', fgColor='FFF2CC'),
    '防雨': PatternFill('solid', fgColor='DEEBF7'),
    '儿童': PatternFill('solid', fgColor='FCE4D6'),
    '电子': PatternFill('solid', fgColor='E4DFEC'),
    '洗漱': PatternFill('solid', fgColor='F2F2F2'),
    '药品': PatternFill('solid', fgColor='FCE4D6'),
    '其他': PatternFill('solid', fgColor='F7F9FC'),
}

BUDGET_ROWS = [
    ('大交通', '上海 → 普吉（3人）', '7/9 春秋；含20kg行李', 2000, 1844, '', '已订'),
    ('大交通', '机票行李/选座补差（普吉段）', '第二笔支付', 1500, 1188, '', '已订'),
    ('大交通', '普吉 → 吉隆坡（3人）', '7/12 早班；亚航', 2000, 1669, '', '已订'),
    ('大交通', '吉隆坡 → 上海（3人）', '7/15 红眼；含税费', 4500, 4100, '', '已订'),
    ('住宿', '普吉芭东万怡 × 3晚', '7/9-7/11；1间含早', 2800, 2500, '', '已订'),
    ('住宿', 'Moxy Chinatown × 2晚', '7/12-7/13；老城', 1100, 900, '', '已订'),
    ('住宿', 'Imperial Lexis × 1晚', '7/14；KLCC无边泳池', 1600, 1400, '', '已订'),
    ('餐饮', '全程餐饮（7天 × 3人）', '酒店早+本地餐馆；少景区餐厅', 3000, None, '', ''),
    ('当地交通', '普吉接送机（往返）', 'Grab 包车', 300, None, '', ''),
    ('当地交通', '吉隆坡 KUL 接送机（往返）', '7/12抵达+7/15离境', 280, None, '', ''),
    ('当地交通', 'Grab + MRT/LRT', '7天市内短途', 650, None, '', ''),
    ('游玩体验', '蛋岛/珊瑚岛半日出海（3人）', '7/10', 900, None, '', ''),
    ('游玩体验', '泰式按摩 × 1次（家庭）', '7/10 晚上', 400, None, '可选', ''),
    ('游玩体验', '双威水上乐园（3人）', '7/13 周一；Klook', 850, None, '', ''),
    ('游玩体验', '双子塔观景台（3人）', '7/14 16:00', 650, None, '', ''),
    ('手续杂项', '签证 / MDAC', '中泰马30天免签；马入境前填MDAC', 0, 0, '', '免费'),
    ('旅行用品', '防水手机袋', '出海+乐园', 50, None, '', ''),
    ('旅行用品', '沙滩拖鞋 / 速干毛巾', '缺则当地买', 150, None, '可选', ''),
    ('旅行用品', '湿巾 / 零食 / 旅行分装', '随身+儿童', 120, None, '', ''),
]

PACKING_ROWS = [
    ('证件', '护照（3人）', '3本', '', '', '有效期6个月+'),
    ('证件', '机票 / 酒店订单（电子版+截图）', '1套', '', '', '含普吉+吉隆坡全程'),
    ('证件', '马来西亚 MDAC 入境卡', '3份', '', '', '出发前72小时内填写'),
    ('证件', '门票二维码截图', '若干', '', '', '双威7/13、双峰塔7/14'),
    ('衣物', '短袖 / 短裤 / 长裤', '各3套', '', '', '7天换洗'),
    ('衣物', '泳衣 + 泳帽（3人）', '3套', '', '', '双威+酒店泳池'),
    ('衣物', '沙滩鞋 / 拖鞋', '3双', '', '', '普吉海滩+乐园'),
    ('衣物', '薄外套或防晒衫', '3件', '', '', '商场冷气+飞机'),
    ('衣物', '一次性内衣裤 / 袜子', '若干', '', '可选', '省心可备'),
    ('防晒', '防晒霜 SPF50+', '2-3支', '', '', '2小时补涂'),
    ('防晒', '遮阳帽', '3顶', '', '', ''),
    ('防晒', '太阳镜', '3副', '', '可选', ''),
    ('防雨', '轻便雨衣 / 折叠伞', '3套', '', '', '7/13午后可能阵雨'),
    ('儿童', '儿童臂圈 / 救生衣', '1套', '', '', '双威+普吉出海'),
    ('儿童', '儿童水杯 / 零食', '各1', '', '', '随身'),
    ('电子', '手机 + 充电器', '3套', '', '', ''),
    ('电子', 'Sony 相机 + 存储卡', '1套', '', '', ''),
    ('电子', '手表 + 充电器', '1套', '', '可选', ''),
    ('电子', '充电宝（20000mAh）', '2个', '', '', '双威日必备'),
    ('电子', '英标插头 Type G', '2-3个', '', '', '新马泰通用'),
    ('电子', '防水手机袋', '3个', '¥50', '', '计入全预算'),
    ('洗漱', '牙刷牙膏 / 旅行装洗护', '1套', '', '', '酒店有备可少带'),
    ('药品', '肠胃药 / 创可贴 / 晕车药', '1盒', '', '', '全家常备'),
    ('药品', '驱蚊液', '1瓶', '', '', '普吉傍晚'),
    ('其他', '湿巾 / 纸巾', '若干', '¥120', '', '计入全预算'),
    ('其他', 'U型颈枕', '3个', '', '可选', '红眼航班'),
    ('其他', '泰铢 + 令吉现金', '少量', '', '', '小摊/储物柜；主卡+Grab'),
    ('其他', 'RM200-300 零钱', '1袋', '', '', '吉隆坡小摊专用'),
]


def style_cell(cell, *, fill=None, font=None, align=None, border=True, fmt=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def apply_header_row(ws, row: int, headers: list[str], ncol: int):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, align=CENTER)
    for c in range(len(headers) + 1, ncol + 1):
        style_cell(ws.cell(row, c), fill=HEADER_FILL, font=HEADER_FONT, align=CENTER)


def clear_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged.bounds
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def format_budget_sheet(ws):
    clear_sheet(ws)
    ncol = 7
    ws.merge_cells('A1:G1')
    t = ws['A1']
    t.value = '7天全程预算（普吉 + 吉隆坡）'
    style_cell(t, fill=HEADER_FILL, font=TITLE_FONT, align=CENTER, border=False)

    ws.merge_cells('A2:G2')
    s = ws['A2']
    s.value = '2大1小 · 7天6晚（7/9普吉→7/12-15吉隆坡）· 金额单位：人民币元 · D列预估上限 / E列实际支付'
    style_cell(s, font=SUB_FONT, align=LEFT, border=False)

    headers = ['分类', '项目', '说明', '预估上限(¥)', '实际支付(¥)', '可选', '是否已订']
    hdr_row = 4
    apply_header_row(ws, hdr_row, headers, ncol)

    start = hdr_row + 1
    for i, row_data in enumerate(BUDGET_ROWS):
        r = start + i
        cat, item, note, est, actual, optional, booked = row_data
        vals = [cat, item, note, est, actual, optional, booked]
        row_fill = CAT_FILL.get(cat, ALT_FILL if i % 2 else PatternFill())
        if i % 2:
            row_fill = ALT_FILL if cat not in CAT_FILL else CAT_FILL[cat]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            fill = CAT_FILL.get(cat) if c <= 3 else (ALT_FILL if i % 2 else PatternFill())
            if c <= 3 and cat in CAT_FILL:
                fill = CAT_FILL[cat]
            elif c > 3:
                fill = ALT_FILL if i % 2 else PatternFill()
            style_cell(cell, fill=fill, font=BODY_FONT, align=LEFT if c <= 3 else (RIGHT if c in (4, 5) else CENTER))
            if c in (4, 5) and isinstance(v, (int, float)):
                cell.number_format = '#,##0'

    end = start + len(BUDGET_ROWS) - 1
    total_row = end + 1
    ws.merge_cells(f'A{total_row}:C{total_row}')
    tc = ws.cell(total_row, 1, '★ 全程预算合计')
    style_cell(tc, fill=TOTAL_FILL, font=TOTAL_FONT, align=LEFT)
    for c in range(2, 4):
        style_cell(ws.cell(total_row, c), fill=TOTAL_FILL, font=TOTAL_FONT, border=True)

    est_cell = ws.cell(total_row, 4)
    est_cell.value = f'=SUM(D{start}:D{end})'
    est_cell.number_format = '#,##0'
    style_cell(est_cell, fill=TOTAL_FILL, font=TOTAL_FONT, align=RIGHT)

    act_cell = ws.cell(total_row, 5)
    act_cell.value = f'=SUM(E{start}:E{end})'
    act_cell.number_format = '#,##0'
    style_cell(act_cell, fill=TOTAL_FILL, font=TOTAL_FONT, align=RIGHT)

    for c in range(6, ncol + 1):
        style_cell(ws.cell(total_row, c), fill=TOTAL_FILL, font=TOTAL_FONT, align=CENTER)

    # 待支付提示行
    tip_row = total_row + 1
    ws.merge_cells(f'A{tip_row}:G{tip_row}')
    tip = ws.cell(tip_row, 1)
    tip.value = f'=CONCATENATE("已支付 ¥",TEXT(E{total_row},"#,##0"),"  ·  预估总上限 ¥",TEXT(D{total_row},"#,##0"),"  ·  待支付约 ¥",TEXT(MAX(0,D{total_row}-E{total_row}),"#,##0"))'
    style_cell(tip, font=Font(name='微软雅黑', size=9, bold=True, color='C55A11'), align=LEFT, border=False)

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[hdr_row].height = 22
    ws.freeze_panes = f'A{hdr_row + 1}'
    ws.auto_filter.ref = f'A{hdr_row}:G{end}'


def format_packing_sheet(ws):
    clear_sheet(ws)
    ncol = 6
    ws.merge_cells('A1:F1')
    t = ws['A1']
    t.value = '出行携带清单'
    style_cell(t, fill=HEADER_FILL, font=TITLE_FONT, align=CENTER, border=False)

    ws.merge_cells('A2:F2')
    s = ws['A2']
    s.value = '出发前逐项勾选 · 标预算项已计入「7天全预算」· 与「行前准备」任务表互补'
    style_cell(s, font=SUB_FONT, align=LEFT, border=False)

    headers = ['分类', '物品', '数量', '预算(¥)', '已准备 ☐', '备注']
    hdr_row = 4
    apply_header_row(ws, hdr_row, headers, ncol)

    start = hdr_row + 1
    for i, row_data in enumerate(PACKING_ROWS):
        r = start + i
        cat, item, qty, budget, done, note = row_data
        vals = [cat, item, qty, budget, done, note]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            fill = CAT_FILL.get(cat) if c == 1 else (ALT_FILL if i % 2 else PatternFill())
            if c == 1 and cat in CAT_FILL:
                fill = CAT_FILL[cat]
            elif c > 1:
                fill = ALT_FILL if i % 2 else PatternFill()
            style_cell(cell, fill=fill, font=BODY_FONT, align=LEFT if c in (2, 6) else CENTER)

    end = start + len(PACKING_ROWS) - 1
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 28
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[hdr_row].height = 22
    ws.freeze_panes = f'A{hdr_row + 1}'
    ws.auto_filter.ref = f'A{hdr_row}:F{end}'


def set_tab_colors(wb):
    from openpyxl.styles.colors import Color

    wb['7天全预算'].sheet_properties.tabColor = Color(rgb='FF375623')
    if '行前准备与携带清单' in wb.sheetnames:
        wb['行前准备与携带清单'].sheet_properties.tabColor = Color(rgb='FF44546A')


def main():
    wb = openpyxl.load_workbook(EXCEL)
    format_budget_sheet(wb['7天全预算'])
    set_tab_colors(wb)
    wb.save(EXCEL)
    print(f'已完善 → {EXCEL}')


if __name__ == '__main__':
    main()
