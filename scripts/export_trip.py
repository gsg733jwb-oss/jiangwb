#!/usr/bin/env python3
"""Export KL travel Excel to web + miniprogram data bundles."""
import argparse
import json
import re
import sys
from datetime import datetime

import openpyxl
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from paths import (  # noqa: E402
    MP_ALL_PLACES,
    MP_ALL_TRIPS,
    MP_DATA_DIR,
    MP_MANIFEST,
    PLACES_JSON,
    WEB_DATA_DIR,
    WEB_DATA_INLINE,
    WEB_DATA_OUT,
    ensure_dirs,
    migrate_legacy_files,
)
from trip_config import (  # noqa: E402
    excel_for_trip,
    load_trip_config,
    resolve_active_trip,
    sync_manifest,
)


def sheet_to_dict(ws, header_row=1):
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            v = ws.cell(r, i + 1).value
            if v is not None and str(v).strip():
                empty = False
            row[str(h)] = v
        if not empty:
            rows.append(row)
    return rows


def export_full_budget(ws):
    title = str(ws.cell(1, 1).value or '').strip()
    subtitle = str(ws.cell(2, 1).value or '').strip()
    rows = sheet_to_dict(ws, header_row=4)
    cleaned = []
    for row in rows:
        cat = row.get('分类')
        item = row.get('项目')
        if not cat and not item:
            continue
        cleaned.append(row)
    return {'title': title, 'subtitle': subtitle, 'rows': cleaned}


def export_checklist(ws):
    title = str(ws.cell(1, 1).value or '').strip()
    subtitle = str(ws.cell(2, 1).value or '').strip()
    rows = sheet_to_dict(ws, header_row=4)
    return {'title': title, 'subtitle': subtitle, 'rows': rows}


def split_type_cat(label):
    text = str(label or '').strip()
    if '·' in text:
        a, b = text.split('·', 1)
        return a.strip(), b.strip()
    return text, '其他'


def derive_food_dist(restaurants):
    rows = []
    for r in restaurants:
        day = str(r.get('安排日') or '').strip()
        if not day or '备选' in day:
            continue
        m = re.match(r'(\d+/\d+)\s*(.*)', day)
        if not m:
            continue
        meal = m.group(2).strip() or '—'
        rows.append({
            '日期': m.group(1),
            '餐次': meal,
            '餐厅': r.get('餐厅'),
            '必点': r.get('必点菜品'),
            '人均RM': r.get('人均(RM)'),
            '说明': r.get('地址'),
        })
    return rows


def export_merged_budget(ws):
    title = str(ws.cell(1, 1).value or '7天全程预算').strip()
    subtitle = str(ws.cell(2, 1).value or '').strip()
    total_budget = ws.cell(5, 6).value
    total_paid = ws.cell(5, 7).value

    budget_rows = []
    full_rows = []
    checklist_rows = []

    for r in range(6, ws.max_row + 1):
        src = ws.cell(r, 1).value
        if not src:
            continue
        cat = ws.cell(r, 2).value
        date = ws.cell(r, 3).value
        item = ws.cell(r, 4).value
        note = ws.cell(r, 5).value
        est = ws.cell(r, 6).value
        paid = ws.cell(r, 7).value
        optional = ws.cell(r, 8).value
        status = ws.cell(r, 9).value

        if src == '预算明细':
            budget_rows.append({
                '日期': date,
                '类别': cat,
                '项目': item,
                '说明': note or '',
                '预算(¥)': est,
                '实际支付(¥)': paid,
                '可选': optional or None,
            })
        elif src == '7天全预算':
            full_rows.append({
                '分类': cat,
                '项目': item,
                '说明': note,
                '预算(¥)': est,
                '实际支付(¥)': paid,
                '可选': optional or None,
                '状态': status or '',
            })
        elif src == '行前准备':
            type_name, sub_cat = split_type_cat(cat)
            checklist_rows.append({
                '类型': type_name,
                '分类': sub_cat,
                '事项/物品': item,
                '何时/数量': date,
                '必备': '是',
                '预算(¥)': est,
                '备注': note,
            })

    if total_budget is not None or total_paid is not None:
        full_rows.append({
            '分类': '合计',
            '项目': None,
            '说明': None,
            '预算(¥)': total_budget,
            '实际支付(¥)': total_paid,
            '可选': None,
            '状态': None,
        })
        if total_budget and total_paid:
            pct = round(float(total_paid) / float(total_budget) * 100, 1)
            full_rows.append({
                '分类': f'已支付约 {pct}%（¥{total_paid:,.0f} / ¥{total_budget:,.0f}）',
                '项目': None,
                '说明': None,
                '预算(¥)': None,
                '实际支付(¥)': None,
                '可选': None,
                '状态': None,
            })

    return (
        budget_rows,
        {'title': title, 'subtitle': subtitle, 'rows': full_rows},
        {
            'title': '行前准备与携带清单',
            'subtitle': '出发前逐项勾选 · 与 Excel「预算」表同步',
            'rows': checklist_rows,
        },
    )


def export_day_sheet(ws):
    headers = []
    for c in range(2, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers.append((c, str(h).strip()))

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        display = ws.cell(r, 1).value
        if display is not None and str(display).strip() != '':
            row['显示'] = int(display) if isinstance(display, (int, float)) else display

        empty = not row
        for c, h in headers:
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                empty = False
            if h in ('坐标起', '坐标落') and isinstance(v, str):
                v = v.strip().rstrip(',，')
            row[h] = v

        if empty:
            continue
        if '显示' not in row:
            row['显示'] = 1
        rows.append(row)
    return rows


def export_food_rankings(ws):
    sections = []
    current = None
    headers = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not a:
            continue
        text = str(a).strip()
        if text.startswith('【') and 'Top' in text:
            current = {'title': text, 'items': []}
            sections.append(current)
            headers = None
            continue
        if text == '排名':
            headers = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            continue
        if headers and isinstance(a, (int, float)):
            item = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                item[str(h)] = ws.cell(r, i + 1).value
            current['items'].append(item)

    return sections


def export_workbook(wb, trip_id: str, meta: dict) -> dict:
    restaurants = sheet_to_dict(wb['美食餐厅'])

    if '预算' in wb.sheetnames:
        budget, full_budget, checklist = export_merged_budget(wb['预算'])
    else:
        budget = sheet_to_dict(wb['预算明细'])
        full_budget = export_full_budget(wb['7天全预算'])
        checklist = export_checklist(wb['行前准备与携带清单'])

    food_dist = (
        sheet_to_dict(wb['美食分布'])
        if '美食分布' in wb.sheetnames
        else derive_food_dist(restaurants)
    )

    out = {
        'id': trip_id,
        'title': meta.get('title') or '出行攻略',
        'subtitle': meta.get('subtitle') or '',
        'timezone': meta.get('timezone') or 'Asia/Shanghai',
        'dateStart': meta.get('dateStart'),
        'dateEnd': meta.get('dateEnd'),
        'overview': sheet_to_dict(wb['行程总览']),
        'days': {},
        'foodDist': food_dist,
        'restaurants': restaurants,
        'foodRankings': export_food_rankings(wb['吉隆坡美食推荐']),
        'mapList': sheet_to_dict(wb['地图清单对照']),
        'budget': budget,
        'fullBudget': full_budget,
        'checklist': checklist,
    }
    for name in wb.sheetnames:
        if name.startswith('Day'):
            out['days'][name.replace('Day', '').strip()] = export_day_sheet(wb[name])
    return out


def _load_json(path: Path) -> dict:
    if path.is_file():
        return json.loads(path.read_text(encoding='utf-8'))
    return {}


def sync_miniprogram_bundle(trip_id: str, trip_data: dict, cfg: dict, updated_at: str) -> None:
    MP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    all_trips = _load_json(MP_ALL_TRIPS)
    all_trips[trip_id] = trip_data
    MP_ALL_TRIPS.write_text(
        json.dumps(all_trips, ensure_ascii=False, indent=2, default=str) + '\n',
        encoding='utf-8',
    )

    places = []
    if PLACES_JSON.is_file():
        places = json.loads(PLACES_JSON.read_text(encoding='utf-8'))
    all_places = _load_json(MP_ALL_PLACES)
    all_places[trip_id] = places
    MP_ALL_PLACES.write_text(
        json.dumps(all_places, ensure_ascii=False, indent=2) + '\n',
        encoding='utf-8',
    )

    meta = cfg['trips'][trip_id]
    manifest = sync_manifest(cfg, trip_id, meta, updated_at)
    MP_MANIFEST.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + '\n',
        encoding='utf-8',
    )


def main():
    parser = argparse.ArgumentParser(description='从 Excel 导出行程到网页与小程序')
    parser.add_argument('--trip-id', help='指定 trip.config 中的行程 id')
    args = parser.parse_args()

    migrate_legacy_files()
    ensure_dirs()
    cfg = load_trip_config()
    trip_id, meta = resolve_active_trip(cfg, args.trip_id)
    excel = excel_for_trip(meta)

    wb = openpyxl.load_workbook(excel, data_only=True)
    out = export_workbook(wb, trip_id, meta)
    updated_at = datetime.now().strftime('%Y-%m-%d %H:%M')

    text = json.dumps(out, ensure_ascii=False, indent=2, default=str)
    inline_js = f'window.__TRIP_DATA__ = {text};\n'
    WEB_DATA_OUT.parent.mkdir(parents=True, exist_ok=True)
    WEB_DATA_OUT.write_text(text, encoding='utf-8')
    WEB_DATA_INLINE.write_text(inline_js, encoding='utf-8')

    stamp = datetime.now().strftime('%Y-%m-%d_%H%M')
    WEB_DATA_DIR.mkdir(parents=True, exist_ok=True)
    (WEB_DATA_DIR / 'trip.json').write_text(text, encoding='utf-8')
    (WEB_DATA_DIR / 'trip.inline.js').write_text(inline_js, encoding='utf-8')
    (WEB_DATA_DIR / f'trip-{stamp}.json').write_text(text, encoding='utf-8')
    (WEB_DATA_DIR / f'{trip_id}-trip-{stamp}.json').write_text(text, encoding='utf-8')

    sync_miniprogram_bundle(trip_id, out, cfg, updated_at)

    print(f'Trip ID  → {trip_id}')
    print(f'Excel    → {excel}')
    print(f'Web      → {WEB_DATA_OUT}')
    print(f'MiniProg → {MP_ALL_TRIPS} (+ manifest)')
    print(
        f'  days={len(out["days"])} checklist={len(out["checklist"]["rows"])} '
        f'fullBudget={len(out["fullBudget"]["rows"])}'
    )


if __name__ == '__main__':
    main()
