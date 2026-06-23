#!/usr/bin/env python3
"""优化 Day 行程表：显示/坐标起/坐标落与全表统一配色、列宽、冻结窗格。"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))
from paths import TRIP_EXCEL, migrate_legacy_files  # noqa: E402

EXCEL = TRIP_EXCEL

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='微软雅黑', size=9)
DISPLAY_MAIN_FONT = Font(name='微软雅黑', size=9, bold=True, color='375623')
DISPLAY_SUB_FONT = Font(name='微软雅黑', size=9, color='808080')
COORD_FONT = Font(name='微软雅黑', size=8, color='44546A')
COORD_MAIN_FONT = Font(name='微软雅黑', size=8, bold=True, color='2F5496')
DIVIDER_FONT = Font(name='微软雅黑', size=9, bold=True, color='1F4E79')
THIN = Side(style='thin', color='D0D7E2')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

TYPE_FILL = {
    '交通': PatternFill('solid', fgColor='DEEBF7'),
    '步行': PatternFill('solid', fgColor='F2F2F2'),
    '购物': PatternFill('solid', fgColor='FFF9E6'),
    '酒店': PatternFill('solid', fgColor='E4DFEC'),
    '餐饮': PatternFill('solid', fgColor='FFF2CC'),
    '乐园': PatternFill('solid', fgColor='D9F2F5'),
    '景点': PatternFill('solid', fgColor='E2EFDA'),
    '项目': PatternFill('solid', fgColor='EDE7F6'),
    '休闲': PatternFill('solid', fgColor='ECEFF1'),
}
ALT_FILL = PatternFill('solid', fgColor='F7F9FC')
DIVIDER_FILL = PatternFill('solid', fgColor='D6E4F0')

COL_WIDTHS = {
    'A': 5, 'B': 16, 'C': 20, 'D': 12, 'E': 28, 'F': 8,
    'G': 32, 'H': 10, 'I': 10, 'J': 10, 'K': 36, 'L': 14, 'M': 28,
}


def _header_map(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }


def _is_divider_row(ws, row: int, headers: dict[str, int]) -> bool:
    activity_col = headers.get('活动/站点')
    if not activity_col:
        return False
    text = str(ws.cell(row, activity_col).value or '')
    return '━━' in text or '════' in text


def _row_fill(ws, row: int, headers: dict[str, int]) -> PatternFill:
    if _is_divider_row(ws, row, headers):
        return DIVIDER_FILL
    type_col = headers.get('类型')
    if not type_col:
        return ALT_FILL
    type_name = str(ws.cell(row, type_col).value or '').strip()
    if not type_name or type_name == '—':
        return ALT_FILL
    return TYPE_FILL.get(type_name, ALT_FILL)


def _display_font(display) -> Font:
    if display in (1, '1', 1.0):
        return DISPLAY_MAIN_FONT
    if display in (2, '2', 2.0):
        return DISPLAY_SUB_FONT
    return BODY_FONT


def _coord_font(display) -> Font:
    if display in (1, '1', 1.0):
        return COORD_MAIN_FONT
    return COORD_FONT


def style_day_sheet(ws):
    headers = _header_map(ws)
    if '显示' not in headers:
        ws.cell(1, 1, '显示')
        headers = _header_map(ws)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        if not cell.value:
            continue
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'D2'

    display_col = headers.get('显示', 1)
    coord_start_col = headers.get('坐标起', 2)
    coord_end_col = headers.get('坐标落', 3)

    for r in range(2, ws.max_row + 1):
        display = ws.cell(r, display_col).value
        divider = _is_divider_row(ws, r, headers)
        fill = _row_fill(ws, r, headers)

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = CENTER if c == display_col else LEFT

            if divider:
                cell.font = DIVIDER_FONT
            elif c == display_col:
                cell.font = _display_font(display)
            elif c in (coord_start_col, coord_end_col):
                cell.font = _coord_font(display) if cell.value else BODY_FONT
            else:
                cell.font = BODY_FONT


def main():
    migrate_legacy_files()
    wb = openpyxl.load_workbook(EXCEL)
    for name in wb.sheetnames:
        if name.startswith('Day'):
            style_day_sheet(wb[name])
    wb.save(EXCEL)
    print(f'已优化 Day 表 → {EXCEL}')


if __name__ == '__main__':
    main()
