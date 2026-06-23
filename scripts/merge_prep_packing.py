#!/usr/bin/env python3
"""合并「行前准备」与「出行全清单」为一张工作表。"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color

sys.path.insert(0, str(Path(__file__).resolve().parent))
from paths import TRIP_EXCEL, migrate_legacy_files  # noqa: E402

EXCEL = TRIP_EXCEL
SHEET_NAME = '行前准备与携带清单'
OLD_SHEETS = ('行前准备', '出行全清单')

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=14)
SUB_FONT = Font(name='微软雅黑', color='44546A', size=9, italic=True)
BODY_FONT = Font(name='微软雅黑', size=9)
ALT_FILL = PatternFill('solid', fgColor='F7F9FC')
TYPE_TASK_FILL = PatternFill('solid', fgColor='E2EFDA')
TYPE_PACK_FILL = PatternFill('solid', fgColor='DEEBF7')
THIN = Side(style='thin', color='D0D7E2')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

HEADERS = ['类型', '分类', '事项/物品', '何时/数量', '必备', '预算(¥)', '完成 ☐', '备注']
NCOL = len(HEADERS)


def style_cell(cell, *, fill=None, font=None, align=None, border=True):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = BORDER


def read_prep_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        item = ws.cell(r, 2).value
        if not item:
            continue
        rows.append([
            '行前任务',
            cat,
            item,
            ws.cell(r, 3).value or '',
            ws.cell(r, 4).value or '',
            '',
            ws.cell(r, 5).value or '',
            ws.cell(r, 6).value or '',
        ])
    return rows


def read_packing_rows(ws):
    rows = []
    for r in range(5, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        item = ws.cell(r, 2).value
        if not item:
            continue
        rows.append([
            '携带物品',
            cat,
            item,
            ws.cell(r, 3).value or '',
            '',
            ws.cell(r, 4).value or '',
            ws.cell(r, 5).value or '',
            ws.cell(r, 6).value or '',
        ])
    return rows


def write_merged_sheet(ws, data_rows):
    ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(NCOL)}1')
    t = ws['A1']
    t.value = '行前准备与携带清单'
    style_cell(t, fill=HEADER_FILL, font=TITLE_FONT, align=CENTER, border=False)

    ws.merge_cells(f'A2:{openpyxl.utils.get_column_letter(NCOL)}2')
    s = ws['A2']
    s.value = '出发前逐项勾选 · 预订/任务与携带物品合一表 · 标预算项见「7天全预算」'
    style_cell(s, font=SUB_FONT, align=LEFT, border=False)

    hdr_row = 4
    for c, h in enumerate(HEADERS, 1):
        style_cell(ws.cell(hdr_row, c, h), fill=HEADER_FILL, font=HEADER_FONT, align=CENTER)

    start = hdr_row + 1
    for i, vals in enumerate(data_rows):
        r = start + i
        row_type = vals[0]
        type_fill = TYPE_TASK_FILL if row_type == '行前任务' else TYPE_PACK_FILL
        for c, v in enumerate(vals, 1):
            fill = type_fill if c == 1 else (ALT_FILL if i % 2 else PatternFill())
            style_cell(ws.cell(r, c, v), fill=fill, font=BODY_FONT, align=LEFT if c in (3, 8) else CENTER)

    end = start + len(data_rows) - 1
    widths = {'A': 10, 'B': 10, 'C': 32, 'D': 14, 'E': 8, 'F': 10, 'G': 9, 'H': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[hdr_row].height = 22
    ws.freeze_panes = f'A{hdr_row + 1}'
    if data_rows:
        ws.auto_filter.ref = f'A{hdr_row}:{openpyxl.utils.get_column_letter(NCOL)}{end}'


def main():
    migrate_legacy_files()
    wb = openpyxl.load_workbook(EXCEL)
    prep_ws = wb['行前准备']
    pack_ws = wb['出行全清单']

    prep_rows = read_prep_rows(prep_ws)
    pack_rows = read_packing_rows(pack_ws)
    merged = prep_rows + pack_rows

    for name in OLD_SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet(SHEET_NAME)
    write_merged_sheet(ws, merged)
    ws.sheet_properties.tabColor = Color(rgb='FF44546A')

    if '预算明细' in wb.sheetnames:
        target = wb.sheetnames.index('预算明细') + 1
        current = wb.sheetnames.index(SHEET_NAME)
        wb.move_sheet(ws, offset=target - current)

    wb.save(EXCEL)
    print(f'已合并 → {EXCEL}')
    print(f'  工作表: {SHEET_NAME}')
    print(f'  行前任务 {len(prep_rows)} 行 + 携带物品 {len(pack_rows)} 行 = {len(merged)} 行')


if __name__ == '__main__':
    main()
