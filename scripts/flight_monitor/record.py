"""桌面记录：Markdown + Excel + JSON。"""

from __future__ import annotations

import json
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ctrip_safari import FlightOffer

PRICE_CHANGE_THRESHOLD = 50
SHEET_NAME = "机票价格监控"
HDR_ROW = 3
COL_COUNT = 11
CONFIG_PATH = Path(__file__).resolve().parent / "config.json"

ROUTE_ORDER = ("sha_hkt", "hkt_kul", "kul_sha")
ROUTE_LABELS = {
    "sha_hkt": "上海 → 普吉岛",
    "hkt_kul": "普吉岛 → 吉隆坡",
    "kul_sha": "吉隆坡 → 上海",
}
LABEL_TO_ID = {v: k for k, v in ROUTE_LABELS.items()}

BOOKED_META = {
    "9C8665": ("春秋航空", "07:45", "12:10", "5小时25分"),
    "AK833": ("马来西亚亚航", "11:10", "13:45", "1小时35分"),
    "9C6524": ("春秋航空", "18:25", "23:55", "5小时30分"),
    "D7330": ("亚航X", "18:40", "00:20", "5小时40分"),
    "AK7330": ("亚航X", "18:40", "00:20", "5小时40分"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="F5F9FC")
UP_FILL = PatternFill("solid", fgColor="FFE5E5")
DOWN_FILL = PatternFill("solid", fgColor="E2F0D9")
NEUTRAL_REMARK_FILL = PatternFill("solid", fgColor="FFF8E5")

THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
SUB_FONT = Font(name="微软雅黑", size=9, color="666666")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
UP_FONT = Font(name="微软雅黑", size=10, bold=True, color="C00000")
DOWN_FONT = Font(name="微软雅黑", size=10, bold=True, color="375623")
NEUTRAL_FONT = Font(name="微软雅黑", size=10, color="7F6000")

HEADERS = [
    "查询时间",
    "航线",
    "航班号",
    "航司",
    "起飞",
    "到达",
    "现价(¥)",
    "上次(¥)",
    "变动(¥)",
    "时长",
    "备注",
]

COLUMN_WIDTHS = {
    "A": 17,
    "B": 18,
    "C": 11,
    "D": 20,
    "E": 9,
    "F": 9,
    "G": 11,
    "H": 11,
    "I": 11,
    "J": 12,
    "K": 42,
}

PRICE_RE = re.compile(r"¥\s*([\d,]+)")
FLIGHT_RE = re.compile(r"\b([A-Z0-9]{2})\s?(\d{3,4})\b")
TIME_RE = re.compile(r"(\d{2}:\d{2})-(\d{2}:\d{2})")
DUR_RE = re.compile(r"(\d+小时\d+分|\d+小时|\d+分)")
FAIL_MARKERS = {"获取失败", "—", "查询失败", None, ""}


@dataclass
class HistoryEntry:
    stamp: str
    route_id: str
    label: str
    flight: FlightOffer | None
    notes: str = ""
    error: str | None = None
    source: str = "history"


def _expand(path: str) -> Path:
    return Path(path).expanduser()


def _normalize_flight_no(value: str) -> str:
    compact = re.sub(r"\s+", "", value.upper())
    match = re.search(r"([A-Z0-9]{2}\d{3,4})", compact)
    return match.group(1) if match else compact


def _load_booked_by_route() -> dict[str, list[str]]:
    booked: dict[str, list[str]] = {
        "kul_sha": ["9C6524", "D7330", "AK7330"],
    }
    if CONFIG_PATH.exists():
        try:
            config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            for route in config.get("routes", []):
                route_id = route.get("id")
                flights = route.get("booked_flights") or []
                if route_id and flights:
                    booked[str(route_id)] = [str(item) for item in flights]
        except (json.JSONDecodeError, OSError):
            pass
    return booked


def _flight_track_key(flight_no: str) -> str:
    normalized = _normalize_flight_no(flight_no)
    if normalized in {"D7330", "AK7330"}:
        return "D7330"
    return normalized


def _matches_booked_target(flight_no: str, targets: set[str]) -> bool:
    normalized = _normalize_flight_no(flight_no)
    if normalized in targets:
        return True
    if normalized in {"D7330", "AK7330"} and ({"D7330", "AK7330"} & targets):
        return True
    return False


def _booked_targets(route_id: str, booked_by_route: dict[str, list[str]]) -> set[str]:
    return {_normalize_flight_no(item) for item in booked_by_route.get(route_id, [])}


def _price_history_key(route_id: str, flight_no: str) -> str:
    return f"{route_id}|{_flight_track_key(flight_no)}"


def _style_cell(
    cell,
    *,
    fill=None,
    font=None,
    align=None,
    border: bool = True,
    number_format: str | None = None,
) -> None:
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = BORDER
    if number_format:
        cell.number_format = number_format


def _wrap_left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center", wrap_text=True)


def _price_delta_note(
    current: int,
    previous: int | None,
    *,
    threshold: int,
) -> tuple[str, PatternFill | None, Font | None]:
    if previous is None:
        return "首次记录", None, NEUTRAL_FONT
    delta = current - previous
    if abs(delta) <= threshold:
        if delta == 0:
            return f"较上次 ¥{previous} 无变化", None, BODY_FONT
        direction = "涨" if delta > 0 else "降"
        return f"较上次 ¥{previous}，{direction} ¥{abs(delta)}（未超 ¥{threshold}）", None, BODY_FONT
    if delta > 0:
        return (
            f"⚠ 涨价：¥{previous} → ¥{current}（+¥{delta}）",
            UP_FILL,
            UP_FONT,
        )
    return (
        f"✓ 降价：¥{previous} → ¥{current}（-¥{abs(delta)}）",
        DOWN_FILL,
        DOWN_FONT,
    )


def _pick_all_booked_flights(
    flights: list[FlightOffer],
    route_id: str,
    booked_by_route: dict[str, list[str]],
) -> list[FlightOffer]:
    if not flights:
        return []
    booked_list = booked_by_route.get(route_id) or []
    if not booked_list:
        return [flights[0]]
    targets = _booked_targets(route_id, booked_by_route)
    matched = [flight for flight in flights if _matches_booked_target(flight.flight_no, targets)]
    if not matched:
        return []
    ordered: list[FlightOffer] = []
    seen_keys: set[str] = set()
    for booked_no in booked_list:
        track_key = _flight_track_key(booked_no)
        if track_key in seen_keys:
            continue
        for flight in matched:
            if _flight_track_key(flight.flight_no) == track_key:
                ordered.append(flight)
                seen_keys.add(track_key)
                break
    return ordered


def _pick_booked_flight(
    flights: list[FlightOffer],
    route_id: str,
    booked_by_route: dict[str, list[str]],
) -> FlightOffer | None:
    picked = _pick_all_booked_flights(flights, route_id, booked_by_route)
    return picked[0] if picked else None


def _parse_kul_sha_from_notes(text: str) -> FlightOffer | None:
    if not text or not re.search(r"D7\s?330|AK7330|D7330", text, re.I):
        return None
    price_match = PRICE_RE.search(text)
    if not price_match:
        return None
    price = int(price_match.group(1).replace(",", ""))
    time_match = TIME_RE.search(text)
    meta = BOOKED_META["D7330"]
    return FlightOffer(
        airline="亚航X",
        flight_no="D7330",
        dep_time=time_match.group(1) if time_match else meta[1],
        arr_time=time_match.group(2) if time_match else meta[2],
        price=price,
        duration=meta[3],
        source="legacy-notes",
    )


def _flight_from_dict(data: dict[str, Any]) -> FlightOffer:
    return FlightOffer(
        airline=str(data.get("airline") or ""),
        flight_no=str(data.get("flight_no") or ""),
        dep_time=str(data.get("dep_time") or ""),
        arr_time=str(data.get("arr_time") or ""),
        price=int(data.get("price") or 0),
        duration=str(data.get("duration") or ""),
        source=str(data.get("source") or ""),
    )


def _parse_legacy_cell(text: Any, route_id: str, booked_by_route: dict[str, list[str]]) -> FlightOffer | None:
    if text in FAIL_MARKERS:
        return None
    raw = str(text).strip()
    if raw in FAIL_MARKERS:
        return None

    per_person = re.search(r"¥\s*([\d,]+)/人", raw)
    if per_person:
        price = int(per_person.group(1).replace(",", ""))
        booked_list = booked_by_route.get(route_id) or []
        booked_no = booked_list[0] if booked_list else ""
        meta = BOOKED_META.get(booked_no, ("—", "—", "—", ""))
        return FlightOffer(
            airline=meta[0],
            flight_no=booked_no or "—",
            dep_time=meta[1],
            arr_time=meta[2],
            price=price,
            duration=meta[3],
            source="legacy",
        )

    price_match = PRICE_RE.search(raw)
    if not price_match:
        return None
    price = int(price_match.group(1).replace(",", ""))

    flight_match = FLIGHT_RE.search(raw)
    flight_no = _normalize_flight_no(flight_match.group(1) + flight_match.group(2)) if flight_match else ""
    targets = _booked_targets(route_id, booked_by_route)
    if targets and flight_no and not _matches_booked_target(flight_no, targets):
        return None
    if not flight_no:
        booked_list = booked_by_route.get(route_id) or []
        flight_no = booked_list[0] if booked_list else ""

    time_match = TIME_RE.search(raw)
    dur_match = DUR_RE.search(raw)
    meta = BOOKED_META.get(flight_no, ("", "", "", ""))

    if flight_match:
        airline = raw.split(flight_match.group(0))[0].strip() or meta[0] or "—"
    else:
        airline = raw.split("¥")[0].strip() or meta[0] or "—"

    dep_time = time_match.group(1) if time_match else meta[1] or "—"
    arr_time = time_match.group(2) if time_match else meta[2] or "—"
    duration = dur_match.group(1) if dur_match else meta[3] or ""

    return FlightOffer(
        airline=airline,
        flight_no=flight_no or "—",
        dep_time=dep_time,
        arr_time=arr_time,
        price=price,
        duration=duration,
        source="legacy",
    )


def _parse_styled_sheet(ws, booked_by_route: dict[str, list[str]]) -> list[HistoryEntry]:
    if ws.max_row <= HDR_ROW:
        return []
    if ws.cell(HDR_ROW, 1).value != "查询时间":
        return []

    entries: list[HistoryEntry] = []
    for row in range(HDR_ROW + 1, ws.max_row + 1):
        stamp = _normalize_stamp(ws.cell(row, 1).value)
        label = str(ws.cell(row, 2).value or "")
        route_id = LABEL_TO_ID.get(label, "")
        if not route_id:
            continue
        price = ws.cell(row, 7).value
        if price in FAIL_MARKERS:
            entries.append(
                HistoryEntry(
                    stamp=stamp,
                    route_id=route_id,
                    label=label,
                    flight=None,
                    notes=str(ws.cell(row, 11).value or ""),
                    error="查询失败",
                    source="sheet",
                )
            )
            continue
        flight_no = str(ws.cell(row, 3).value or "")
        entries.append(
            HistoryEntry(
                stamp=stamp,
                route_id=route_id,
                label=label,
                flight=FlightOffer(
                    airline=str(ws.cell(row, 4).value or ""),
                    flight_no=flight_no,
                    dep_time=str(ws.cell(row, 5).value or ""),
                    arr_time=str(ws.cell(row, 6).value or ""),
                    price=int(price),
                    duration=str(ws.cell(row, 10).value or ""),
                    source="sheet",
                ),
                notes=str(ws.cell(row, 11).value or ""),
                source="sheet",
            )
        )
    return entries


def _save_legacy_archive(json_log_dir: Path, entries: list[HistoryEntry]) -> None:
    if not entries:
        return
    archive = json_log_dir / "_legacy_archive.json"
    payload = [
        {
            "stamp": e.stamp,
            "route_id": e.route_id,
            "label": e.label,
            "notes": e.notes,
            "error": e.error,
            "source": e.source,
            "flight": {
                "airline": e.flight.airline,
                "flight_no": e.flight.flight_no,
                "dep_time": e.flight.dep_time,
                "arr_time": e.flight.arr_time,
                "price": e.flight.price,
                "duration": e.flight.duration,
                "source": e.flight.source,
            }
            if e.flight
            else None,
        }
        for e in entries
    ]
    archive.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_legacy_archive(json_log_dir: Path) -> list[HistoryEntry]:
    archive = json_log_dir / "_legacy_archive.json"
    if not archive.exists():
        return []
    try:
        payload = json.loads(archive.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    entries: list[HistoryEntry] = []
    for item in payload:
        flight_data = item.get("flight")
        flight = _flight_from_dict(flight_data) if flight_data else None
        entries.append(
            HistoryEntry(
                stamp=item.get("stamp", ""),
                route_id=item.get("route_id", ""),
                label=item.get("label", ""),
                flight=flight,
                notes=str(item.get("notes") or ""),
                error=item.get("error"),
                source="legacy",
            )
        )
    return entries


def _normalize_stamp(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s*\(约\)\s*$", "", text)
    if len(text) == 16:
        return text
    if len(text) == 10:
        return text
    return text


def _is_legacy_sheet(ws) -> bool:
    if ws.max_row < 1:
        return False
    first = [ws.cell(1, c).value for c in range(1, 6)]
    return first[:5] == ["时间", "上海→普吉岛", "普吉岛→吉隆坡", "吉隆坡→上海", "备注"]


def _parse_legacy_workbook(xlsx_path: Path, booked_by_route: dict[str, list[str]]) -> list[HistoryEntry]:
    if not xlsx_path.exists():
        return []

    wb = load_workbook(xlsx_path, data_only=True)
    entries: list[HistoryEntry] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not _is_legacy_sheet(ws):
            continue
        for row in range(2, ws.max_row + 1):
            stamp = _normalize_stamp(ws.cell(row, 1).value)
            notes = str(ws.cell(row, 5).value or "").strip()
            route_cells = [
                ("sha_hkt", _parse_legacy_cell(ws.cell(row, 2).value, "sha_hkt", booked_by_route)),
                ("hkt_kul", _parse_legacy_cell(ws.cell(row, 3).value, "hkt_kul", booked_by_route)),
                ("kul_sha", _parse_legacy_cell(ws.cell(row, 4).value, "kul_sha", booked_by_route)),
            ]
            kul_targets = _booked_targets("kul_sha", booked_by_route)
            kul_cell = str(ws.cell(row, 4).value or "")
            kul_flights: list[FlightOffer] = []
            kul_cell_flight = route_cells[2][1]
            if kul_cell_flight and (
                not kul_targets
                or _matches_booked_target(kul_cell_flight.flight_no, kul_targets)
            ):
                kul_flights.append(kul_cell_flight)
            d7330 = _parse_kul_sha_from_notes(notes) or _parse_kul_sha_from_notes(kul_cell)
            if d7330 and not any(
                _flight_track_key(item.flight_no) == "D7330" for item in kul_flights
            ):
                kul_flights.append(d7330)
            route_cells = route_cells[:2] + [("kul_sha", flight) for flight in kul_flights]
            if not any(route_id == "kul_sha" for route_id, _ in route_cells):
                route_cells.append(("kul_sha", None))
            for route_id, flight in route_cells:
                label = ROUTE_LABELS[route_id]
                if flight is None:
                    raw_cell = ws.cell(row, ROUTE_ORDER.index(route_id) + 2).value
                    if raw_cell in FAIL_MARKERS:
                        entries.append(
                            HistoryEntry(
                                stamp=stamp,
                                route_id=route_id,
                                label=label,
                                flight=None,
                                notes=notes if route_id == "kul_sha" else "",
                                error="查询失败",
                                source="legacy",
                            )
                        )
                    continue
                entries.append(
                    HistoryEntry(
                        stamp=stamp,
                        route_id=route_id,
                        label=label,
                        flight=flight,
                        notes=notes if route_id == "kul_sha" else "",
                        source="legacy",
                    )
                )
    return entries


def _load_json_entries(json_log_dir: Path, booked_by_route: dict[str, list[str]]) -> list[HistoryEntry]:
    entries: list[HistoryEntry] = []
    if not json_log_dir.exists():
        return entries

    runs: list[dict[str, Any]] = []
    for path in sorted(json_log_dir.glob("*.json")):
        if path.name.startswith("_"):
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        if isinstance(data.get("runs"), list):
            runs.extend(data["runs"])
        elif data.get("run_at"):
            runs.append(data)

    for run in runs:
        run_at = run.get("run_at", "")
        try:
            stamp = datetime.fromisoformat(run_at).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            stamp = _normalize_stamp(run_at)
        notes = str(run.get("notes") or "").strip()
        for route in run.get("routes", []):
            route_id = route.get("id") or LABEL_TO_ID.get(route.get("label", ""), "")
            if not route_id:
                continue
            label = route.get("label") or ROUTE_LABELS.get(route_id, route_id)
            flights = [_flight_from_dict(item) for item in route.get("flights") or []]
            picked = _pick_all_booked_flights(flights, route_id, booked_by_route)
            if picked:
                for flight in picked:
                    entries.append(
                        HistoryEntry(
                            stamp=stamp,
                            route_id=route_id,
                            label=label,
                            flight=flight,
                            notes=notes if route_id == "kul_sha" else "",
                            source="json",
                        )
                    )
            elif route.get("error") or not flights:
                entries.append(
                    HistoryEntry(
                        stamp=stamp,
                        route_id=route_id,
                        label=label,
                        flight=None,
                        notes=notes if route_id == "kul_sha" else "",
                        error=str(route.get("error") or "查询失败"),
                        source="json",
                    )
                )
    return entries


def _normalize_history_entry(
    entry: HistoryEntry,
    booked_by_route: dict[str, list[str]],
) -> HistoryEntry | None:
    if entry.flight is None:
        return entry
    targets = _booked_targets(entry.route_id, booked_by_route)
    if not targets:
        return entry
    if _matches_booked_target(entry.flight.flight_no, targets):
        return entry
    if entry.route_id == "kul_sha":
        alt = _parse_kul_sha_from_notes(entry.notes)
        if alt:
            return HistoryEntry(
                stamp=entry.stamp,
                route_id=entry.route_id,
                label=entry.label,
                flight=alt,
                notes=entry.notes,
                error=entry.error,
                source=entry.source,
            )
    return None


def _collect_history_entries(xlsx_path: Path, json_log_dir: Path) -> list[HistoryEntry]:
    booked_by_route = _load_booked_by_route()
    legacy = _parse_legacy_workbook(xlsx_path, booked_by_route)
    if legacy:
        _save_legacy_archive(json_log_dir, legacy)
    else:
        legacy = _load_legacy_archive(json_log_dir)
        if not legacy and xlsx_path.exists():
            wb = load_workbook(xlsx_path, data_only=True)
            if SHEET_NAME in wb.sheetnames:
                legacy = _parse_styled_sheet(wb[SHEET_NAME], booked_by_route)
                if legacy:
                    _save_legacy_archive(json_log_dir, legacy)
    json_entries = _load_json_entries(json_log_dir, booked_by_route)
    merged = _merge_history_entries(legacy, json_entries)
    normalized: list[HistoryEntry] = []
    for entry in merged:
        item = _normalize_history_entry(entry, booked_by_route)
        if item is not None:
            normalized.append(item)
    return normalized


def _entry_sort_key(entry: HistoryEntry) -> tuple[str, int, str]:
    route_idx = ROUTE_ORDER.index(entry.route_id) if entry.route_id in ROUTE_ORDER else 9
    flight_key = _flight_track_key(entry.flight.flight_no) if entry.flight else "ZZZ"
    return entry.stamp, route_idx, flight_key


def _entry_dedupe_key(entry: HistoryEntry) -> str:
    flight_key = _flight_track_key(entry.flight.flight_no) if entry.flight else "ERR"
    return f"{entry.stamp}|{entry.route_id}|{flight_key}"


def _entry_priority(entry: HistoryEntry, booked_by_route: dict[str, list[str]]) -> tuple[int, int, int]:
    source_rank = 0 if entry.source == "json" else 1
    booked_rank = 1
    targets = _booked_targets(entry.route_id, booked_by_route)
    if entry.flight and targets:
        if _matches_booked_target(entry.flight.flight_no, targets):
            booked_rank = 0
    error_rank = 1 if entry.flight else 0
    return source_rank, booked_rank, error_rank


def _merge_history_entries(*groups: list[HistoryEntry]) -> list[HistoryEntry]:
    booked_by_route = _load_booked_by_route()
    merged: dict[str, HistoryEntry] = {}
    for group in groups:
        for entry in group:
            key = _entry_dedupe_key(entry)
            existing = merged.get(key)
            if existing is None or _entry_priority(entry, booked_by_route) < _entry_priority(
                existing, booked_by_route
            ):
                merged[key] = entry
    return sorted(merged.values(), key=_entry_sort_key)


def _setup_sheet(ws, *, threshold: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
    title = ws.cell(1, 1, "机票价格监控 · 已订航班跟踪")
    _style_cell(title, fill=TITLE_FILL, font=TITLE_FONT, align=_center(), border=False)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COL_COUNT)
    subtitle = ws.cell(
        2,
        1,
        f"7/9 9C8665 · 7/12 AK833 · 7/15 9C6524 + D7330  |  备注列：变动超过 ¥{threshold} 以颜色标注",
    )
    _style_cell(subtitle, font=SUB_FONT, align=_wrap_left(), border=False)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(HDR_ROW, col, header)
        _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, align=_center())

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[HDR_ROW].height = 24
    ws.freeze_panes = f"A{HDR_ROW + 1}"


def _write_history_row(
    ws,
    row: int,
    entry: HistoryEntry,
    *,
    previous_price: int | None,
    threshold: int,
    zebra: bool,
) -> None:
    row_fill = ALT_FILL if zebra else PatternFill()
    if entry.flight:
        flight = entry.flight
        current_price = int(flight.price)
        remark, remark_fill, remark_font = _price_delta_note(
            current_price,
            previous_price,
            threshold=threshold,
        )
        if entry.notes:
            remark = f"{remark}  |  {entry.notes}"
        delta = "" if previous_price is None else current_price - previous_price
        values: list[Any] = [
            entry.stamp,
            entry.label,
            flight.flight_no,
            flight.airline,
            flight.dep_time,
            flight.arr_time,
            current_price,
            previous_price if previous_price is not None else "—",
            delta if delta != "" else "—",
            flight.duration or "—",
            remark,
        ]
    else:
        err = entry.error or "查询失败"
        remark = f"{err}  |  {entry.notes}".strip(" |")
        values = [
            entry.stamp,
            entry.label,
            "—",
            "—",
            "—",
            "—",
            "—",
            previous_price if previous_price is not None else "—",
            "—",
            "—",
            remark,
        ]
        remark_fill = NEUTRAL_REMARK_FILL
        remark_font = UP_FONT

    for col, value in enumerate(values, start=1):
        cell = ws.cell(row, col, value)
        align = _wrap_left() if col in (2, 4, 11) else _center()
        if col in (7, 8, 9) and isinstance(value, (int, float)):
            align = _right()
        font = BODY_FONT
        fill = row_fill
        if col == 11:
            fill = remark_fill or row_fill
            font = remark_font or BODY_FONT
        elif col == 9 and isinstance(value, int):
            if value > threshold:
                fill = UP_FILL
                font = UP_FONT
            elif value < -threshold:
                fill = DOWN_FILL
                font = DOWN_FONT
        _style_cell(
            cell,
            fill=fill,
            font=font,
            align=align,
            number_format="#,##0" if col in (7, 8) and isinstance(value, int) else None,
        )
    ws.row_dimensions[row].height = 42


def rebuild_excel_workbook(
    xlsx_path: Path,
    json_log_dir: Path,
    *,
    threshold: int = PRICE_CHANGE_THRESHOLD,
) -> int:
    entries = _collect_history_entries(xlsx_path, json_log_dir)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _setup_sheet(ws, threshold=threshold)

    last_prices: dict[str, int] = {}
    row = HDR_ROW + 1
    for idx, entry in enumerate(entries):
        price_key = (
            _price_history_key(entry.route_id, entry.flight.flight_no)
            if entry.flight
            else entry.route_id
        )
        prev = last_prices.get(price_key)
        _write_history_row(
            ws,
            row,
            entry,
            previous_price=prev,
            threshold=threshold,
            zebra=idx % 2 == 1,
        )
        if entry.flight:
            last_prices[price_key] = int(entry.flight.price)
        row += 1

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return len(entries)


def _append_json_log(json_path: Path, payload: dict[str, Any]) -> None:
    runs: list[dict[str, Any]] = []
    if json_path.exists():
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
            if isinstance(data.get("runs"), list):
                runs = data["runs"]
            elif data.get("run_at"):
                runs = [data]
        except (json.JSONDecodeError, OSError):
            runs = []
    runs.append(payload)
    json_path.write_text(
        json.dumps({"runs": runs}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _load_price_history(json_log_dir: Path) -> dict[str, int]:
    history: dict[str, int] = {}
    booked_by_route = _load_booked_by_route()
    for entry in _load_json_entries(json_log_dir, booked_by_route):
        if entry.flight:
            history[_price_history_key(entry.route_id, entry.flight.flight_no)] = int(
                entry.flight.price
            )
    return history


def _maybe_migrate_outputs(
    desktop: Path,
    json_dir: Path,
    *,
    excel_file: str,
    markdown_file: str,
) -> None:
    """首次换目录时，从桌面旧位置复制已有记录。"""
    old_desktop = Path.home() / "Desktop"
    old_flight_dir = old_desktop / "机票监控"
    old_json = old_desktop / "机票监控日志"
    guide_flight = old_desktop / "马泰攻略" / "机票监控"
    desktop.mkdir(parents=True, exist_ok=True)
    json_dir.mkdir(parents=True, exist_ok=True)

    sources = [
        old_desktop,
        old_flight_dir,
        guide_flight,
    ]
    for name in (excel_file, markdown_file):
        dst = desktop / name
        if dst.exists():
            continue
        for src_dir in sources:
            src = src_dir / name
            if src.exists():
                shutil.copy2(src, dst)
                break

    if not any(json_dir.glob("*.json")):
        for src_dir in (old_json, old_flight_dir / "日志", guide_flight / "日志"):
            if not src_dir.is_dir():
                continue
            for path in src_dir.glob("*.json"):
                shutil.copy2(path, json_dir / path.name)
            if any(json_dir.glob("*.json")):
                break


def save_records(
    *,
    desktop_dir: str,
    markdown_file: str,
    excel_file: str,
    json_log_dir: str,
    run_at: datetime,
    route_results: list[dict[str, Any]],
    notes: str = "",
    price_change_threshold: int = PRICE_CHANGE_THRESHOLD,
) -> dict[str, str]:
    desktop = _expand(desktop_dir)
    json_dir = _expand(json_log_dir)
    _maybe_migrate_outputs(
        desktop,
        json_dir,
        excel_file=excel_file,
        markdown_file=markdown_file,
    )
    desktop.mkdir(parents=True, exist_ok=True)
    json_dir.mkdir(parents=True, exist_ok=True)

    stamp = run_at.strftime("%Y-%m-%d %H:%M")
    day_key = run_at.strftime("%Y-%m-%d")
    json_path = json_dir / f"{day_key}.json"

    last_prices = _load_price_history(json_dir)

    serializable = {
        "run_at": run_at.isoformat(timespec="seconds"),
        "notes": notes,
        "routes": [
            {
                "id": item.get("id"),
                "label": item["label"],
                "date": item["date"],
                "source": item.get("source", ""),
                "direct_count": item.get("direct_count"),
                "flights": [
                    {
                        "airline": f.airline,
                        "flight_no": f.flight_no,
                        "dep_time": f.dep_time,
                        "arr_time": f.arr_time,
                        "price": f.price,
                        "duration": f.duration,
                        "source": f.source,
                    }
                    for f in item.get("flights", [])
                ],
                "error": item.get("error"),
                "scrape_hint": item.get("scrape_hint"),
            }
            for item in route_results
        ],
    }
    _append_json_log(json_path, serializable)

    md_path = desktop / markdown_file
    md_block = _build_markdown_block(
        stamp,
        route_results,
        notes,
        last_prices=last_prices,
        threshold=price_change_threshold,
    )
    if md_path.exists():
        existing = md_path.read_text(encoding="utf-8")
        md_path.write_text(md_block + "\n\n---\n\n" + existing, encoding="utf-8")
    else:
        header = (
            "# 机票监控记录\n\n"
            "> 每日跟踪已订航班：7/9 07:45 9C8665 上海→普吉、"
            "7/12 11:10 AK833 普吉→吉隆坡、7/15 9C6524 + D7330 吉隆坡→上海。\n"
            f"> 变动超过 ¥{price_change_threshold} 会在备注标红/标绿。\n\n"
        )
        md_path.write_text(header + md_block + "\n", encoding="utf-8")

    xlsx_path = desktop / excel_file
    row_count = rebuild_excel_workbook(
        xlsx_path,
        json_dir,
        threshold=price_change_threshold,
    )

    return {
        "markdown": str(md_path),
        "excel": str(xlsx_path),
        "json": str(json_path),
        "excel_rows": str(row_count),
    }


def _build_markdown_block(
    stamp: str,
    route_results: list[dict[str, Any]],
    notes: str,
    *,
    last_prices: dict[str, int],
    threshold: int,
) -> str:
    lines = [f"## {stamp}", ""]
    if notes:
        lines.extend([f"> {notes}", ""])
    for item in route_results:
        lines.append(f"### {item['label']}（{item['date']}）")
        source = item.get("source")
        if source:
            lines.append(f"- 数据源：{source}")
        if item.get("direct_count") is not None:
            lines.append(f"- 直飞航班数：{item['direct_count']}")
        if item.get("error"):
            lines.append(f"- 错误：{item['error']}")
        flights: list[FlightOffer] = item.get("flights", [])
        if flights:
            lines.append("")
            lines.append("| # | 航司 | 航班号 | 起飞-到达 | 现价 | 变动 | 时长 | 备注 |")
            lines.append("|---:|---|---|---|---:|---:|---|---|")
            route_id = str(item.get("id") or "")
            for idx, flight in enumerate(flights, start=1):
                current = int(flight.price)
                price_key = _price_history_key(route_id, flight.flight_no) if route_id else ""
                previous = last_prices.get(price_key) if price_key else None
                if previous is None:
                    delta_text = "—"
                    remark = "首次记录"
                else:
                    delta = current - previous
                    delta_text = f"{delta:+d}" if delta else "0"
                    remark, _, _ = _price_delta_note(current, previous, threshold=threshold)
                lines.append(
                    f"| {idx} | {flight.airline} | {flight.flight_no} | "
                    f"{flight.dep_time}-{flight.arr_time} | ¥{current} | {delta_text} | "
                    f"{flight.duration or '—'} | {remark} |"
                )
        else:
            lines.append("")
            lines.append("暂无直飞结果。")
        lines.append("")
    return "\n".join(lines).rstrip()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="重建机票监控 Excel（合并旧记录）")
    parser.add_argument("--desktop", default="~/Desktop/马泰攻略/机票监控")
    parser.add_argument("--excel", default="机票监控.xlsx")
    parser.add_argument("--json-log-dir", default="~/Desktop/马泰攻略/机票监控/日志")
    parser.add_argument("--threshold", type=int, default=PRICE_CHANGE_THRESHOLD)
    args = parser.parse_args()
    desktop = _expand(args.desktop)
    json_dir = _expand(args.json_log_dir)
    _maybe_migrate_outputs(
        desktop,
        json_dir,
        excel_file=args.excel,
        markdown_file="机票监控记录.md",
    )
    count = rebuild_excel_workbook(
        desktop / args.excel,
        json_dir,
        threshold=args.threshold,
    )
    print(f"已重建 {desktop / args.excel}，共 {count} 行")
